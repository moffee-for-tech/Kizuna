import os
import re
import uuid
import logging
from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from slowapi import Limiter
from slowapi.util import get_remote_address

from middleware.rbac import get_current_user
from config import settings

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/upload", tags=["upload"])
limiter = Limiter(key_func=get_remote_address)

ALLOWED_EXTENSIONS = {".pdf", ".docx", ".csv", ".xlsx", ".txt", ".md"}
ALLOWED_MIME_TYPES = {
    ".pdf": {"application/pdf"},
    ".docx": {"application/vnd.openxmlformats-officedocument.wordprocessingml.document"},
    ".csv": {"text/csv", "application/csv", "text/plain"},
    ".xlsx": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    ".txt": {"text/plain"},
    ".md": {"text/plain", "text/markdown"},
}
# Magic byte signatures for binary file types
FILE_SIGNATURES = {
    ".pdf": [b"%PDF"],
    ".docx": [b"PK\x03\x04"],
    ".xlsx": [b"PK\x03\x04"],
}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB


def extract_text(file_path: str, extension: str) -> str:
    """Extract text from uploaded file."""
    try:
        if extension == ".pdf":
            from PyPDF2 import PdfReader
            reader = PdfReader(file_path)
            return "\n".join(page.extract_text() or "" for page in reader.pages)

        elif extension == ".docx":
            from docx import Document
            doc = Document(file_path)
            return "\n".join(para.text for para in doc.paragraphs)

        elif extension in (".csv", ".txt", ".md"):
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()

        elif extension == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            text_parts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    text_parts.append(
                        "\t".join(str(cell) if cell else "" for cell in row)
                    )
            return "\n".join(text_parts)

        return ""

    except Exception as e:
        logger.error(f"Text extraction failed for {file_path}: {e}")
        return ""


@router.post("")
@limiter.limit("10/minute")
async def upload_file(
    request: Request,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """Upload a document for context injection."""
    # Validate extension
    _, ext = os.path.splitext(file.filename or "")
    ext = ext.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type: {ext}. Allowed: {', '.join(ALLOWED_EXTENSIONS)}",
        )

    # Validate MIME type
    content_type = file.content_type or ""
    allowed_mimes = ALLOWED_MIME_TYPES.get(ext, set())
    if allowed_mimes and content_type and content_type not in allowed_mimes:
        raise HTTPException(
            status_code=400,
            detail=f"MIME type '{content_type}' does not match extension '{ext}'",
        )

    # Read and check size
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File too large. Max 10MB.")

    # Validate file magic bytes for binary types
    signatures = FILE_SIGNATURES.get(ext)
    if signatures and not any(content.startswith(sig) for sig in signatures):
        raise HTTPException(
            status_code=400,
            detail=f"File content does not match expected format for '{ext}'",
        )

    # Sanitize filename to prevent path traversal
    original_name = file.filename or "upload"
    safe_name = re.sub(r'[^\w\-.]', '_', os.path.basename(original_name))
    unique_name = f"{uuid.uuid4().hex[:12]}_{safe_name}"

    # Save file
    upload_dir = os.path.join(settings.UPLOAD_DIR, current_user["department"])
    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, unique_name)

    # Verify resolved path stays within upload directory
    if not os.path.realpath(file_path).startswith(os.path.realpath(upload_dir)):
        raise HTTPException(status_code=400, detail="Invalid file path")

    with open(file_path, "wb") as f:
        f.write(content)

    # Extract text
    text = extract_text(file_path, ext)
    word_count = len(text.split()) if text else 0

    return {
        "filename": file.filename,
        "size": len(content),
        "type": ext,
        "word_count": word_count,
        "text_preview": text[:500] + "..." if len(text) > 500 else text,
        "text": text,
    }
